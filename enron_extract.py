from nltk.corpus import stopwords
import re
import openpyxl
from os import listdir, walk
from os.path import isfile, join

wb = openpyxl.Workbook()
wb.create_sheet("Training", index=0)
training_sheet = wb.get_sheet_by_name("Training")
training_sheet.cell(row=1, column=1).value = "CATEGORY"
training_sheet.cell(row=1, column=2).value = "MESSAGE"
feature_list =[]


ham_count = spam_count = count = 2
STOP_WORDS = list(set(stopwords.words('english')))
STOP_WORDS.append('subject:')


def remove_stop_words(word_list):
    word_list = filter(lambda word:word not in STOP_WORDS and len(word) >= 4, word_list)
    return list(word_list)

def filter_line(line):
    line = re.sub('((www\.[^\s]+)|(https?://[^\s]+))','',line)
    line = re.sub("[^a-zA-Z/\s=!-\"\"]+"," ", line)
    line = re.sub('[\s]+', ' ', line)
    line = line.strip('\'"')
    return line
def pre_process_ham(file_name):
    f = open(file_name, 'r')
    mail = " ".join(str(i.lower()) for i in f)
    mail = mail.splitlines()
    subj_in = body_in = 0
    for i in range(len(mail)):
        if "subject" in mail[i]:
            subj = mail[i]
        elif "filename" in mail[i]:
            body_in = i+2
            break
    word_list = []
    word_list.extend(subj.split())
    for i in range(body_in, len(mail)):
        line = filter_line(mail[i])
        '''
        line = re.sub('((www\.[^\s]+)|(https?://[^\s]+))','',line)
        line = re.sub("[^a-zA-Z/\s=!-\"\"]+"," ", line)
        line = re.sub('[\s]+', ' ', line)
        line = line.strip('\'"')
        '''
        word_list.extend(line.split())
    word_list = remove_stop_words(word_list)
    line = " ".join(str(word) for word in word_list)
    return line


def add_to_xl(path, tag):
    global count, spam_count, ham_count
    file_names = [f for f in listdir(path) if isfile(join(path, f))]
    for file_name in file_names:
        try:
            if tag == "ham":
                line = pre_process_ham(path+"/"+file_name)
            else:
                word_list = []
                line = ""
                for _line in open(path+file_name, 'r'):
                    #line_t = filter_line(_line)
                    word_list.extend(_line.lower().split())
                    #line += line_t
                word_list = remove_stop_words(word_list)
                line = " ".join(str(word) for word in word_list)
                line = re.sub('((www\.[^\s]+)|(https?://[^\s]+))','',line)
                line = re.sub("[^a-zA-Z/\s=!-\"\"]+"," ", line)
            if len(line.split()) < 25:
                    continue
            else:
                if tag == 'ham':
                    training_sheet.cell(row=count, column=1).value = "ham"
                    ham_count += 1
                else:
                    training_sheet.cell(row=count, column=1).value = "spam"
                    spam_count += 1
            training_sheet.cell(row=count, column=2).value = line
            feature_list.extend(line.split())
            count += 1
            #if count > 2000:
             #   break
            print(count)
        except:
            continue


def make_xl():
    path_ham = "C:/Users/arpit/Desktop/raw_data/ham/"
    path_spam = 'C:/Users/arpit/Desktop/raw_data/spam/'
    parent = [x[0] for x in walk(path_ham)]
    c=0
    for path in parent:
        add_to_xl(path, 'ham')
        #c+=1
        #if c>70:
         #   break
    add_to_xl(path_spam, 'spam')

make_xl()
print(count, ham_count, spam_count)
f = open("feature_list.txt", 'w')
feature_list = sorted(list(set(feature_list)))
print(len(feature_list))
'''
for i in range(len(feature_list)):
    word = re.sub("[^a-zA-Z/\s=!-\"\"]+","", feature_list[i])
    if len(word) <= 3:
        del feature_list[i]
'''
feature_list = list(filter(lambda word:len(re.sub("[^a-zA-Z/\s=!-\"\"]+","", word)) > 3, feature_list))
feature_list = list(map(lambda word:re.sub("[^a-zA-Z/\s=!-\"\"]+","", word), feature_list))
print(len(feature_list))
for feature in feature_list:
    f.write(feature+"\n")
f.close()
print(feature)
wb.save("processed_mails.xlsx")