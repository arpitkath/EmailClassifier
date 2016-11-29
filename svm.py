import numpy as np
from sklearn import svm
import openpyxl
from random import randint
feature_list = [feature.strip() for feature in open("feature_list.txt", 'r')]


def get_feature_list(word_list):
    feature_vec = []
    dict = {}
    for feature in feature_list:
        dict[feature] = 0
    for word in word_list:
        if word in dict:
            dict[word] = 1
    return list(dict.values())


def get_feature_vec():
    feature_vector = []
    labels = []
    wb = openpyxl.load_workbook("processed_mails.xlsx")
    training_sheet = wb.get_sheet_by_name("Training")
    c=0
    for i in range(1, training_sheet.get_highest_row()):
        try:
            label = training_sheet.cell(row=i, column=1).value
            word_list = training_sheet.cell(row=i, column=2).value.split()
            feature = get_feature_list(word_list)
            feature_vector.append(feature)
            if label == 'ham':
                labels.append(0)
            else:
                labels.append(1)
            print(i)
        except:
            continue
    return feature_vector, labels

feature_vector, labels = get_feature_vec()
p = (80*len(feature_vector))//100
count=0
test_vector = []
test_labels = []
while count < p:
    r = randint(0, len(feature_vector)-1)
    test_vector.append(feature_vector[r])
    test_labels.append(labels[r])
    del feature_vector[r]
    del labels[r]
    count+=1

#feature_vector = np.array(feature_vector)
print("training started")
cl = svm.SVC(kernel='linear', C=1.0)
print("continue")
cl.fit(feature_vector, labels)
print("training end")
print(cl.score(test_vector, test_labels))
