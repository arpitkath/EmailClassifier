import nltk.classify
import openpyxl
from nltk.corpus import stopwords
from random import randint
from collections import Counter
import re

wb = openpyxl.load_workbook('processed_mails.xlsx')


training_sheet = wb.get_sheet_by_name("Training")

training_set = []
test_set = []
count = 0

def prepare():
    for i in range(2, training_sheet.max_row):
        topic = training_sheet.cell(row=i, column=1).value
        body = training_sheet.cell(row=i, column=2).value
        body = dict(Counter(body))
        tup = (body, topic)
        training_set.append(tup)
prepare()
TEST_COUNT = 30*len(training_set)//100
while len(test_set) < TEST_COUNT:
    t = randint(0, len(training_set)-1)
    test_set.append(training_set[t])
    del training_set[t]
print(len(training_set), len(test_set))
cl = nltk.classify.NaiveBayesClassifier.train(training_set)
print("Accuracy of the classifier: {0}".format(nltk.classify.accuracy(cl, test_set)))
cl.show_most_informative_features()
s = input("Input:")
while len(s)>0:
    print(cl.classify(dict(Counter(s))))
    s = input("Input:")